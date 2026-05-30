"""
XLSX converter for MarkItDown-Pro

Note: This converter requires openpyxl to be installed.
Install with: pip install openpyxl
"""

from typing import Dict, Any, Optional, List
from .base import BaseConverter


class XLSXConverter(BaseConverter):
    """
    Converter for XLSX/XLS files
    
    Converts Excel spreadsheets to Markdown with multiple sheet support.
    Requires openpyxl: pip install openpyxl
    """
    
    name = "xlsx"
    description = "XLSX to Markdown converter"
    supported_extensions = ['.xlsx', '.xls']
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize XLSX converter
        
        Args:
            config: Optional configuration
        """
        super().__init__(config)
        
        # Try to import openpyxl
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX conversion. "
                "Install with: pip install openpyxl"
            )
    
    def convert(self, file_path: str, options: Optional[Dict[str, Any]] = None) -> str:
        """
        Convert XLSX file to Markdown
        
        Args:
            file_path: Path to XLSX file
            options: Conversion options
            
        Returns:
            Markdown content
        """
        options = options or {}
        include_empty = options.get('include_empty', False)
        max_rows = options.get('max_rows', 1000)
        sheet_names = options.get('sheets', None)  # Specific sheets to convert
        
        lines = []
        lines.append("# Excel Spreadsheet\n")
        
        try:
            # Load workbook
            wb = self.openpyxl.load_workbook(file_path, data_only=True)
            
            lines.append(f"**Sheets**: {', '.join(wb.sheetnames)}\n")
            
            # Process sheets
            sheets_to_process = sheet_names if sheet_names else wb.sheetnames
            
            for sheet_name in sheets_to_process:
                if sheet_name not in wb.sheetnames:
                    continue
                
                sheet = wb[sheet_name]
                
                lines.append(f"## Sheet: {sheet_name}\n")
                
                # Get sheet dimensions
                max_row = min(sheet.max_row, max_rows)
                max_col = sheet.max_column
                
                if max_row == 0 or max_col == 0:
                    lines.append("*Empty sheet*\n")
                    continue
                
                lines.append(f"**Dimensions**: {max_row} rows × {max_col} columns\n")
                
                # Extract data
                rows = []
                for row_idx in range(1, max_row + 1):
                    row_data = []
                    has_data = False
                    
                    for col_idx in range(1, max_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        if value is not None:
                            has_data = True
                            # Format value
                            if isinstance(value, (int, float)):
                                value = str(value)
                            else:
                                value = str(value).strip()
                                # Truncate long strings
                                if len(value) > 100:
                                    value = value[:100] + "..."
                        else:
                            value = ""
                        
                        row_data.append(value)
                    
                    if has_data or include_empty:
                        rows.append(row_data)
                
                # Create table
                if rows:
                    # Use first row as headers if it looks like headers
                    headers = rows[0]
                    data_rows = rows[1:]
                    
                    lines.append(self.create_table(headers, data_rows))
                    lines.append("")
                
                if sheet.max_row > max_rows:
                    lines.append(f"*{sheet.max_row - max_rows} more rows not shown*\n")
            
            wb.close()
            
        except Exception as e:
            lines.append(f"**Error converting XLSX**: {str(e)}")
        
        return '\n'.join(lines)
    
    def get_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata from XLSX file
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            Dictionary of metadata
        """
        metadata = super().get_metadata(file_path)
        
        try:
            wb = self.openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            metadata['sheet_count'] = len(wb.sheetnames)
            metadata['sheet_names'] = wb.sheetnames
            
            # Get dimensions for each sheet
            sheet_info = {}
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_info[sheet_name] = {
                    'rows': sheet.max_row,
                    'columns': sheet.max_column,
                }
            
            metadata['sheets'] = sheet_info
            
            wb.close()
            
        except Exception:
            pass
        
        return metadata
